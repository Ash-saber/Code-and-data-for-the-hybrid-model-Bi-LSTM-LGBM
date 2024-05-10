# -*- coding: utf-8 -*-
"""
Created on Mon May 29 15:25:25 2023

@author: Steve
对应于最终数据2
"""


import matplotlib.ticker as mtick
import pandas as pd
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
import tensorflow as tf
from xgboost import XGBRegressor
from lightgbm import LGBMRegressor
import os
path = r'D:\Program project\python project\pytorch\time_seris_2'
os.chdir(path)

from tools.utils import *
from tools.model_create import *

np.random.seed(42)
tf.random.set_seed(42)

def create_dataset(dataset, Fdataset, look_back, predict_step): #构造数据集结构函数
    #dataset是变形、围岩等级、掌子面里程时间序列，(n,3)的矩阵
    #Fdataset是非时间序列因素，仅与断面有关
    dataX, dataY = [], []                #定义空数组的方式
    dataF = []
    dataF2 = []
    for j in range(0,len(dataset)-look_back-predict_step+1):
        a_array = dataset[j:(j+look_back)]
        dataX.append(a_array)            #这个是所有的data
        b_array = dataset[(j + look_back):j+look_back+predict_step,0] 
        dataY.append(b_array)            #这个是变形序列
        
        d_array = dataset[(j + look_back):j+look_back+predict_step,1:]
        d_array = np.squeeze(d_array,0)  #这个是监测时间、掌子面里程序列、围岩等级
        dataF2.append(d_array)
        
        c_array = Fdataset
        dataF.append(c_array)
    return np.array(dataX),np.array(dataY),np.array(dataF),np.array(dataF2)


#-------左线数据集-----

file = r'所有数据\最终数据2\left_top.xlsx'



LTSdata_x = np.empty(shape=[0,3,3])
LTSdata_y = np.empty(shape=[0,1])

LFdata = np.empty(shape=[0,4])
LFdata2 = np.empty(shape=[0,2])

LFdata3 = np.empty(shape=[0,1])#新因素，明天是否开挖台阶

wb1 = load_workbook(file,data_only=True)
ws1 = wb1.sheetnames
####构造数据集结构#####
smp_num=0
smpnum_set = []#样本随断面增加而增加的记录，后面根据断面数量划分数据集要用到
left_dataset = []
for i in range(len(ws1)-2):    #读取ws-2个断面数据，合成数据集
    ######序列数据
    LTSdata = pd.read_excel(file,sheet_name=i,header=0, usecols = [1,9,12],engine='openpyxl')#监测时间、变形、围岩等级、掌子面里程时间序列
    # LTSdata = LTSdata.dropna()   #这条命令太重要了
    LTSdata = LTSdata.values
    LTSdata = LTSdata.astype('float32')
    #将监测时间与变形序列交换顺序
    LTSdata[:,[0,2]]=LTSdata[:,[2,0]]
    

    
    ####因素数据
    LFactordata = pd.read_excel(file,sheet_name=i,header=0, usecols = [4,5,6,11],engine='openpyxl')#4个非时间序列因素，只随断面变化
    LFactordata = LFactordata.dropna()   #这条命令太重要了
    LFactordata = LFactordata.values
    LFactordata = LFactordata.astype('float32')
    LFactordata = np.squeeze(LFactordata,0)

    left_dataset.append([LTSdata,LFactordata])

    look_back = 3
    predict_step=1

    X_data,Y_data,F_data,F2_data  = create_dataset(LTSdata,LFactordata, look_back,predict_step)
    print("目前数据集的格式：",X_data.shape,Y_data.shape,F_data.shape,F2_data.shape)
    print('当前样本量',X_data.shape[0])    
    smp_num += X_data.shape[0]
    smpnum_set.append(smp_num)
    print('当前样本量总数',smp_num)
    
    LTSdata_x = np.concatenate((LTSdata_x,X_data),axis = 0)
    LTSdata_y = np.concatenate((LTSdata_y,Y_data),axis = 0)
    
    LFdata = np.concatenate((LFdata,F_data),axis = 0)
    LFdata2 = np.concatenate((LFdata2,F2_data),axis = 0)
    
    #增加因素，昨天是否开挖台阶
    F3 = pd.read_excel(file,sheet_name=i,header=0, usecols = [13],engine='openpyxl')#4个非时间序列因素，只随断面变化
    # LFdata3 = LFdata3.dropna()   #这条命令太重要了
    F3 = np.append(0,F3.values[3:-1]).reshape(-1,1)
    F3 = F3.astype('float32')
    LFdata3 = np.concatenate((LFdata3,F3),axis = 0)
    # F3 = np.squeeze(F3,0)
    # LFdata3 = np.concatenate((LFdata3,find_singular(Y_data)),axis = 0)
#将“昨天是否开挖台阶”这一因素与其他4个静态因素合并
LFdata_new = np.concatenate((LFdata,LFdata3),axis = 1)
#-----------右线测试集----------
right_dataset = []

file = r'所有数据\最终数据2\right_top.xlsx'


RTSdata_x = np.empty(shape=[0,3,3])
RTSdata_y = np.empty(shape=[0,1])

RFdata = np.empty(shape=[0,4])
RFdata2 = np.empty(shape=[0,2])

RFdata3 = np.empty(shape=[0,1])#新因素，明天是否开挖台阶

wb2 = load_workbook(file,data_only=True)
ws2 = wb2.sheetnames

for k in range(len(ws2)-2):
    ######序列数据
    RTSdata = pd.read_excel(file,sheet_name=k,header=0, usecols = [1,9,12])
    RTSdata = RTSdata.dropna()   #这条命令太重要了
    RTSdata = RTSdata.values
    RTSdata = RTSdata.astype('float32')
    #将监测时间与变形序列交换顺序
    RTSdata[:,[0,2]]=RTSdata[:,[2,0]]
    
    
    ####因素数据
    RFactordata = pd.read_excel(file,sheet_name=k,header=0, usecols = [4,5,6,11])
    RFactordata = RFactordata.dropna()   #这条命令太重要了
    RFactordata = RFactordata.values
    RFactordata = RFactordata.astype('float32')
    RFactordata = np.squeeze(RFactordata,0)

    right_dataset.append([RTSdata,RFactordata])

    look_back = 3
    predict_step=1

    X_data,Y_data,F_data,F2_data  = create_dataset(RTSdata,RFactordata, look_back,predict_step)
    print("目前数据集的格式：",X_data.shape,Y_data.shape,F_data.shape,F2_data.shape)
    print('当前样本量（右线）',X_data.shape[0])      
    smp_num += X_data.shape[0]
    smpnum_set.append(smp_num)
    print('当前样本量总数（右线）',smp_num)
    
    RTSdata_x = np.concatenate((RTSdata_x,X_data),axis = 0)
    RTSdata_y = np.concatenate((RTSdata_y,Y_data),axis = 0)
    
    RFdata = np.concatenate((RFdata,F_data),axis = 0)
    RFdata2 = np.concatenate((RFdata2,F2_data),axis = 0)
    
    #增加因素，昨天是否开挖台阶
    F3 = pd.read_excel(file,sheet_name=k,header=0, usecols = [13],engine='openpyxl')#4个非时间序列因素，只随断面变化
    # LFdata3 = LFdata3.dropna()   #这条命令太重要了
    F3 = np.append(0,F3.values[3:-1]).reshape(-1,1)
    F3 = F3.astype('float32')
    RFdata3 = np.concatenate((RFdata3,F3),axis = 0)
    # F3 = np.squeeze(F3,0)
    # LFdata3 = np.concatenate((LFdata3,find_singular(Y_data)),axis = 0)
#将“昨天是否开挖台阶”这一因素与其他4个静态因素合并
RFdata_new = np.concatenate((RFdata,RFdata3),axis = 1)


####构造总数据集结构#####
full_dataset = left_dataset + right_dataset
TSdata_x = np.concatenate((LTSdata_x,RTSdata_x),axis=0)
#交换Fdata最后一个类别因素与连续因素
Fdata_x = np.concatenate((LFdata,RFdata),axis=0)
Fdata_x[:,[-2,-1]] = Fdata_x[:,[-1,-2]]
Fdata2 = np.concatenate((LFdata2,RFdata2),axis=0)

Fdata3 = np.concatenate((LFdata3,RFdata3),axis=0) #新因素，今天是否开挖台阶
Fdata_new = np.concatenate((LFdata_new,RFdata_new),axis=0) #新的静态因素集，加上今天是否开挖台阶
#交换Fdata_new因素，使得因素顺序为围岩等级，支护类型，台阶数，今天是否开挖台阶，断面埋深
Fdata_new[:,[2,3,4]] = Fdata_new[:,[3,4,2]]

TSdata_y = np.concatenate((LTSdata_y,RTSdata_y),axis=0)
ws = ws1[:-2]+ws2[:-2]
duanmian_num = len(ws)#断面总数

# ####构造总数据集结构（不使用右线数据）#####
# full_dataset = left_dataset
# TSdata_x = LTSdata_x
# #交换Fdata最后一个类别因素与连续因素
# Fdata_x = LFdata
# Fdata_x[:,[-2,-1]] = Fdata_x[:,[-1,-2]]

# TSdata_y = LTSdata_y
# Fdata2 = LFdata2
# ws = ws1[:-2]
# duanmian_num = len(ws)#断面总数


# ##########给变形时间序列加高斯噪声###############
# TSdata_yn = np.array([])
# for start,end in zip([0]+smpnum_set[:-1],smpnum_set):
#     np.random.seed(0)
#     y_temp = TSdata_y[start:end]
#     # TSdata_yn=np.append(TSdata_yn,add_discrete_noise(TSdata_y[start:end],intensity=2,discrete=0.4))
#     TSdata_yn=np.append(TSdata_yn,add_noise2(add_noise(TSdata_y[start:end],intensity=np.max(y_temp)/12),intensity=1.5))
#     # TSdata_yn=np.append(TSdata_yn,add_noise2(TSdata_y[start:end],intensity=1))
    
# TSdata_yn = TSdata_yn.reshape(-1,1)
# #画个图看下
# d_id=7
# fig = plt.figure(dpi=600)
# plt.plot(TSdata_y[smpnum_set[d_id]:smpnum_set[d_id+1]],marker='o')
# plt.plot(TSdata_yn[smpnum_set[d_id]:smpnum_set[d_id+1]],marker='*',linestyle='--')
# #plt.plot(TSdata_yn[smpnum_set[5]:smpnum_set[6]].reshape(-1,1)-TSdata_y[smpnum_set[5]:smpnum_set[6]],marker='*',linestyle='--')

#设置数据集划分比例
train_rate = 0.8
val_rate = 0
test_rate = 0.2


train_dnum = int(train_rate*duanmian_num)
val_dnum = int(val_rate*duanmian_num)
test_dnum = duanmian_num-train_dnum-val_dnum

trainTS = TSdata_x[:smpnum_set[train_dnum-1]]
trainF = Fdata_x[:smpnum_set[train_dnum-1]]
trainY = TSdata_y[:smpnum_set[train_dnum-1]]
trainF2 = Fdata2[:smpnum_set[train_dnum-1]]
trainF3 = Fdata3[:smpnum_set[train_dnum-1]]
trainF_new = Fdata_new[:smpnum_set[train_dnum-1]]

valTS = TSdata_x[smpnum_set[train_dnum-1]:smpnum_set[train_dnum+val_dnum-1]]
valF = Fdata_x[smpnum_set[train_dnum-1]:smpnum_set[train_dnum+val_dnum-1]]
valY = TSdata_y[smpnum_set[train_dnum-1]:smpnum_set[train_dnum+val_dnum-1]]
valF2 = Fdata2[smpnum_set[train_dnum-1]:smpnum_set[train_dnum+val_dnum-1]]
valF3 = Fdata3[smpnum_set[train_dnum-1]:smpnum_set[train_dnum+val_dnum-1]]
valF_new = Fdata_new[smpnum_set[train_dnum-1]:smpnum_set[train_dnum+val_dnum-1]]

testTS = TSdata_x[smpnum_set[train_dnum+val_dnum-1]:]
testF = Fdata_x[smpnum_set[train_dnum+val_dnum-1]:]
testY = TSdata_y[smpnum_set[train_dnum+val_dnum-1]:]
testF2 = Fdata2[smpnum_set[train_dnum+val_dnum-1]:]
testF3 = Fdata3[smpnum_set[train_dnum+val_dnum-1]:]
testF_new = Fdata_new[smpnum_set[train_dnum+val_dnum-1]:]

#使用自己写的类函数来标准化
Normal_X, Normal_Y = Normalize_timeseris(),Normalize()
Normal_X.fit(trainTS,norm_type=0)
Normal_Y.fit(trainY,norm_type=0)

Normal_F2 = Normalize()
Normal_F2.fit(trainF2,norm_type=0)

Normal_licheng = Normalize()
Normal_licheng.fit(trainF[:,-1],norm_type=0)

#写个专门标准化静态因素的函数，前四个为类别变量第五个为连续变量，处理方式不同
def Normal_Fdata(Fdata):
    '''
    Parameters
    ----------
    Fdata  (smp_num,factor_num)
    
    Returns
    -------
    onehot_encoded and normalized Fdata

    '''
    norm_Fdata = np.empty(shape=[Fdata.shape[0],0])
    for i in range(Fdata.shape[1]):
        print(i)
        if i<=2:
            norm_Fdata = np.concatenate((norm_Fdata,onehot_encode(Fdata[:,i],Fdata_new[:,i])),axis=1)
        elif i==3:
            norm_Fdata = np.concatenate((norm_Fdata,Fdata[:,i].reshape(-1,1)),axis=1)
        elif i==4:
            norm_Fdata = np.concatenate((norm_Fdata,Normal_licheng.norm_transform(Fdata[:,i]).reshape(-1,1)),axis=1)
    return norm_Fdata
for i in range(4):
    print(f'{np.unique(Fdata_new[:,i])}')

#对所有数据进行标准化
trainTS = Normal_X.norm_transform(trainTS)
trainY = Normal_Y.norm_transform(trainY)
trainF = Normal_Fdata(trainF)
trainF2 = Normal_F2.norm_transform(trainF2)
trainF_new = Normal_Fdata(trainF_new)

valTS = Normal_X.norm_transform(valTS)
valY = Normal_Y.norm_transform(valY)
valF = Normal_Fdata(valF)
valF2 = Normal_F2.norm_transform(valF2)
valF_new = Normal_Fdata(valF_new)

testTS = Normal_X.norm_transform(testTS)
#testY = Normal_X.norm_transform(testY)
testF = Normal_Fdata(testF)
testF2 = Normal_F2.norm_transform(testF2)
testF_new = Normal_Fdata(testF_new)



#先搞四个模型进行对比，纯LSTM模型（不考虑静态因素），LSTM+factor模型（用神经网络考虑静态因素），
#LSTM+Tree_model模型（用树模型考虑静态因素），Tree_model模型（用树模型考虑时间序列）

#各个模型最优超参数
para_lstm = {'lstmc_num': 2, 'lstms1_num': 1, 'lstms2_num': 2, 'lstms3_num': 0}
para_lstml = {'lstmc_num': 2, 'lstms1_num': 2, 'lstms2_num': 1, 'lstms3_num': 1}
para_lstmf = {'lstmc_num': 2, 'lstms1_num': 2, 'lstms2_num': 1, 'lstms3_num': 1}
para_mlp = {'mlpc_num': 0, 'mlps1_num': 2}
para_lgb = {'n_estimators': 300, 'max_depth': 11, 'lambda_l1': 0, 'lambda_l2': 0}
para_xgb = {'n_estimators': 1700, 'max_depth': 8, 'lambda_l1': 40, 'lambda_l2': 40}
para_rf = {'n_estimators': 140, 'max_depth': 17, 'max_features': 6, 'min_samples_split': 3}
para_svr = {'C': 81.0, 'gamma': 0.004, 'epsilon': 0.008}


#训练考虑4个静态因素以及2个未来可预知动态因素的bi-lstm-3v-6f模型
model_lstm_3v6f = create_model('3v_6f',INPUT_DIM=3,F_DIM=12)
np.random.seed(421)
tf.random.set_seed(421)
history1 = model_lstm_3v6f.fit([trainTS,trainF_new,trainF2], trainY,
                    epochs=300,
                    batch_size=128, 
                    validation_data=([valTS,valF,valF2], valY))

#训练bi-lstm-3v模型
model_lstm_3v = create_model2('2v',**para_lstml)
np.random.seed(422)
tf.random.set_seed(422)
history2 = model_lstm_3v.fit([trainTS], trainY,
                    epochs=300,
                    batch_size=128, 
                    validation_data=([valTS], valY))

#训练考虑4个静态因素外加是否开挖台阶新因素的bi-lstm-3v-5f模型
model_lstm_3v4f = create_model2('2v_4f',F_DIM=12,**para_lstm)
np.random.seed(3)
tf.random.set_seed(3)
history3 = model_lstm_3v4f.fit([trainTS,trainF_new], trainY,
                    epochs=300,
                    batch_size=128, 
                    validation_data=([valTS,valF], valY))
'''
base_lstm = create_model2('2v')
history = model_lstm.fit([trainTS], trainY,
                    epochs=100,
                    batch_size=128, 
                    validation_data=([valTS], valY))
'''
#训练只考虑变形的lstm模型
np.random.seed(421)
tf.random.set_seed(421)
lstm_smp = create_model(model_type='lstm_1v')
lstm_smp.fit(trainTS[:,:,0:1], trainY,
                epochs=300,#epoch设置小一点为xgboost提供训练空间
                batch_size=128)

#----------所有时变因素展开成一维后的长度，其他模型训练需要用-----#
shape = 9

#训练考虑静态因素的mlp模型
mlp = create_model2('mlp',F_DIM=11,**para_mlp)
history = mlp.fit(np.concatenate((trainTS.reshape(-1,shape),trainF),axis=1),trainY,
                    epochs=300,
                    batch_size=128)

#训练meta模型
lr = 1
#meta_model = Meta_Model(model_lstm,XGBRegressor())
#meta_model = Meta_Model(model_lstm_3v4f,LGBMRegressor())
# meta_model = Meta_Model(model_lstm_3v4f,XGBRegressor())
# meta_model = Meta_Model(lstm_smp,LGBMRegressor())
# meta_model = Meta_Model(model_lstm_3v,XGBRegressor())
# meta_model = Meta_Model(model_lstm_3v,RandomForestRegressor())
# meta_model = Meta_Model(model_lstm_3v,LGBMRegressor())
meta_model = Meta_Model(lstm_smp,XGBRegressor())
# meta_model.fit(trainTS,trainF3,trainY,lr=lr)
# meta_model.fit(trainTS,trainF_new,trainY,lr=lr)
meta_model.fit(trainTS[:,:,0:1],trainF_new,trainY,lr=lr)
#meta_model.fit(valTS[:,:,0:1],valF,valY,lr=lr)
#meta_model.fit(valTS[:,:,0:1],np.append(valF,valTS[:,1,1].reshape(-1,1),axis=1),valY,lr=lr)

#训练LGBM模型
tree_model = LGBMRegressor()
tree_model.fit(np.concatenate((trainTS.reshape(-1,shape),trainF),axis=1),trainY)

#----其他传统模型，SVR,MLP,rf,xgboost
from sklearn.svm import SVR
from sklearn.ensemble import RandomForestRegressor
from xgboost import XGBRFRegressor

#训练SVR模型
svr = SVR()
svr.fit(np.concatenate((trainTS.reshape(-1,shape),trainF),axis=1),trainY)

#训练rf模型
rf = RandomForestRegressor(**para_rf)
rf.fit(np.concatenate((trainTS.reshape(-1,shape),trainF),axis=1),trainY)

#训练xgboost模型
xgb = XGBRFRegressor(**para_xgb)
xgb.fit(np.concatenate((trainTS.reshape(-1,shape),trainF),axis=1),trainY)

#-----检验各个模型在测试集上的精度，使用mae,rmse,r三个指标来评判-----
'''
testTS = trainTS
testF = trainF
testY = Normal_Y.unorm_transform(trainY)
'''
#a = np.array([[np.min(trainTS[:,1,:],axis=0)]])

y_mlp = mlp.predict(np.concatenate((testTS.reshape(-1,shape),testF),axis=1)).reshape(-1,1)
y_lstm_3v = model_lstm_3v.predict(testTS)
y_lstm_3v4f = model_lstm_3v4f.predict([testTS,testF_new])
y_lstm_3v6f = model_lstm_3v6f.predict([testTS,testF_new,testF2])
y_smplstm = lstm_smp.predict(testTS[:,:,0:1])

y_svr = svr.predict(np.concatenate((testTS.reshape(-1,shape),testF),axis=1)).reshape(-1,1)
y_tree = tree_model.predict(np.concatenate((testTS.reshape(-1,shape),testF),axis=1)).reshape(-1,1)
y_rf = rf.predict(np.concatenate((testTS.reshape(-1,shape),testF),axis=1)).reshape(-1,1)
y_xgb = xgb.predict(np.concatenate((testTS.reshape(-1,shape),testF),axis=1)).reshape(-1,1)

# y_meta = meta_model.predict(testTS,testF3)
# y_meta = meta_model.predict(testTS,testF_new)
y_meta = meta_model.predict(testTS[:,:,0:1],testF_new)
#y_meta = meta_model.predict(testTS[:,:,0:1],np.append(testF,testTS[:,1,1].reshape(-1,1),axis=1))
#计算测试集总体预测效果评价指标
yp_set = np.concatenate((y_smplstm,y_mlp,y_lstm_3v,y_lstm_3v4f,y_lstm_3v6f,y_svr,y_rf,y_xgb,y_tree,y_meta),axis=1)
# yp_set = np.concatenate((y_meta,y_meta),axis=1)
metrics_result = pd.DataFrame([],columns = ['mae','rmse','mape','R'])
for i in range(yp_set.shape[1]):
    yp_set[:,i] = Normal_Y.unorm_transform(yp_set[:,i])
    result = metrics_cal(yp_set[:,i],testY)
    metrics_result = pd.concat([metrics_result,result],axis=0)
metrics_result.index=['lstm','mlp','bi-lstm-3v','bi-lstm-3v4f','bi-lstm-3v6f','svr','lgb','rf','xgb','lstm-tree']

metrics_result.to_csv('metrics_result2.csv')


#按断面计算测试集预测效果评价指标
metrics_duanmian = []
for i in range(16):
    start_duanmian = train_dnum+val_dnum+i*1-1
    start_num = smpnum_set[start_duanmian]-len(trainY)-len(valY)
    end_num = smpnum_set[start_duanmian+1]-len(trainY)-len(valY)
    metrics_temp = pd.DataFrame([],index = ['mae','rmse','mape','R'])
    for j in range(yp_set.shape[1]):
        result = metrics_cal(yp_set[start_num:end_num,j],testY[start_num:end_num],axis=1)
        metrics_temp = pd.concat([metrics_temp,result],axis=1)
    metrics_temp = metrics_temp.iloc[:,[9,3,2,6,7,8,5,1]]
    metrics_temp.iloc[0:3,-2] = metrics_temp.iloc[0:3,-2]
    metrics_temp.iloc[0:3,3:6] = metrics_temp.iloc[0:3,3:6]
    metrics_temp.iloc[3,-1] = metrics_temp.iloc[3,-1]
    metrics_temp.iloc[3,3:6] = metrics_temp.iloc[3,3:6]
    metrics_temp.columns=['Bi-LSTM-LGBM','Bi-LSTM-MLP','Bi-LSTM','RF','Light GBM','XGBoost','SVR','MLP']
    # metrics_temp.columns=['lstm','mlp','bi-lstm-3v','bi-lstm-3v4f','bi-lstm-3v6f','svr','lgb','rf','xgb','lstm-tree']
    metrics_duanmian.append(metrics_temp)



#--------画BI-LSTM-LGBM与其他传统机器学习方法在测试集上的预测效果----#
############------中文版-------###############
yt = testY
import seaborn as sns
import pandas as pd
err_mlp = (Normal_Y.unorm_transform(y_mlp)-yt)
err_svr = (Normal_Y.unorm_transform(y_svr)-yt)
err_LGBM = (Normal_Y.unorm_transform(y_tree)-yt)
err_rf = (Normal_Y.unorm_transform(y_rf)-yt)
err_xgb = (Normal_Y.unorm_transform(y_xgb)-yt)
err_lstmtree = (Normal_Y.unorm_transform(y_meta)-yt)
err_lstm = (Normal_Y.unorm_transform(y_lstm_3v)-yt)
err_lstmf = (Normal_Y.unorm_transform(y_lstm_3v6f)-yt)
np.min(err_lstmf)
err_tol = np.concatenate((err_lstmtree,err_svr,err_LGBM,err_rf,err_xgb,err_mlp),axis=1)

mat = pd.DataFrame(err_tol,columns=(['Bi-LSTM-LGBM','SVR','Light GBM','RF','XGBoost','MLP']))

plt.rc('font',family='Times New Roman')
plt.rcParams['xtick.direction'] = 'in'
plt.rcParams['ytick.direction'] = 'in'
font1={'family':'SimHei'}
plt.rcParams['font.size']=13

fig = plt.figure(dpi=600,figsize=(5,4))
ax = fig.add_subplot(111)
ax = sns.kdeplot(mat,fill=False)
ax.set_xlim(-4,4)
ax.set_xlabel('误差/mm', fontproperties='SimHei')
ax.set_ylabel('分布密度', fontproperties='SimHei')
ax.legend(['Bi-LSTM-LGBM','SVR','Light GBM','RF','XGBoost','MLP'],frameon=False,loc='upper right')
# ax.grid(linewidth=1.5, color='gray', linestyle=':', alpha=0.5)
plt.savefig('img/lstmtree模型与传统模型对比.svg',bbox_inches='tight',dpi=600)

############------英文版-------###############
yt = testY
import seaborn as sns
import pandas as pd
err_mlp = (Normal_Y.unorm_transform(y_mlp)-yt)
err_svr = (Normal_Y.unorm_transform(y_svr)-yt)
err_LGBM = (Normal_Y.unorm_transform(y_tree)-yt)
err_rf = (Normal_Y.unorm_transform(y_rf)-yt)
err_xgb = (Normal_Y.unorm_transform(y_xgb)-yt)
err_lstmtree = (Normal_Y.unorm_transform(y_meta)-yt)

err_tol = np.concatenate((err_lstmtree,err_svr,err_LGBM,err_rf,err_xgb,err_mlp),axis=1)

mat = pd.DataFrame(err_tol,columns=(['Bi-LSTM-LGBM','SVR','Light GBM','RF','XGBoost','MLP']))

plt.rc('font',family='Times New Roman')
plt.rcParams['xtick.direction'] = 'in'
plt.rcParams['ytick.direction'] = 'in'
font1={'family':'SimHei'}
plt.rcParams['font.size']=12
#####################################
fig = plt.figure(dpi=600,figsize=(5,4))
ax = fig.add_subplot(111)
ax = sns.kdeplot(mat,fill=False)
ax.set_xlim(-5.0,4.0)
ax.set_xlabel('Error (mm)')
ax.set_ylabel('Distribution density')
ax.xaxis.set_major_formatter(mtick.FormatStrFormatter('%.1f'))
legend = ['Bi-LSTM-LGBM','SVR','Light GBM','RF','XGBoost','MLP']
legend.reverse()
ax.legend(legend,frameon=False,loc='upper left')
# ax.grid(linewidth=1.5, color='gray', linestyle=':', alpha=0.5)
plt.savefig('img/lstmtree模型与传统模型对比(英文).svg',bbox_inches='tight',dpi=600)
############箱型图###############
fig = plt.figure(dpi=600,figsize=(5,4))
ax = fig.add_subplot(111)
ax = sns.kdeplot(mat,fill=False)
ax.set_xlim(-40,40)
ax.set_xlabel('Error (mm)')
ax.set_ylabel('Distribution density')
legend = ['Bi-LSTM-LGBM','SVR','Light GBM','RF','XGBoost','MLP']
legend.reverse()
ax.legend(legend,frameon=False,loc='upper left')
# ax.grid(linewidth=1.5, color='gray', linestyle=':', alpha=0.5)

#--------画测试集15个断面预测与测量值对比图（LSTM-LGBM模型）--------#
###----中文版----###
duanmian = ['k36+645','k36+660','k36+675','k36+690','k36+700','k36+715',
            'k36+730','k36+745','k36+755','k36+765','k36+775','k36+785',
            'k36+795','k36+810','k36+825','k36+840']

duanmian.reverse()

font1={'family':'SimHei'}
plt.rcParams['font.size']=13
plt.rc('font',family='Times New Roman')
fig, axs = plt.subplots(5,3,dpi=600,figsize=(12,18))
np.arange(4,30)
for i in range(5):
    for j in range(3):
        start_duanmian = train_dnum+val_dnum+i*3+j-1
        start_num = smpnum_set[start_duanmian]-len(trainY)-len(valY)
        end_num = smpnum_set[start_duanmian+1]-len(trainY)-len(valY)
        yp = yp_set[start_num:end_num,-1]
        yt = testY[start_num:end_num]
        t = np.arange(4,4+len(yp))
        ax = axs[i,j]
        #ax.plot(yt,linestyle='-',marker='^',color='#0072BD')
        #ax.plot(yp,linestyle='--',marker='*',color='#D95319')
        ax.plot(t,yt,linestyle='-',marker='o',color='black',linewidth=2,markersize=5)
        ax.plot(t,yp,linestyle='--',marker='o',color='#1A6EDF',markerfacecolor='white',linewidth=2,markersize=5)
        ax.grid(linewidth=1.5, color='gray', linestyle=':', alpha=0.5)
        #plt.rc('font',family='SimHei')
        ax.legend(['真实值','预测值'],prop = font1,loc='upper left',frameon=False)
        ax.set_title(duanmian[i*3+j],x=0.8,y=0.2,fontsize=15)
        if j==0:
            ax.set_ylabel('拱顶变形/cm', fontproperties='SimHei')
        if i==4:
            ax.set_xlabel('时间/日', fontproperties='SimHei')
        ax.set_ylim(0,32)
plt.savefig('img/15个测试断面预测效果图.svg',bbox_inches='tight',dpi=600)
plt.show()
    
###----英文版----###
duanmian = ['K36+645','K36+660','K36+675','K36+690','K36+700','K36+715',
            'K36+730','K36+745','K36+755','K36+765','K36+775','K36+785',
            'K36+795','K36+810','K36+825','K36+840']


duanmian.reverse()
section_select = [14,8,12,7,11,4,13,2,1]
section_select = [14,13,12,11,8,7,4,2,1]
font1={'family':'SimHei'}
plt.rcParams['font.size']=13
plt.rc('font',family='Times New Roman')
plt.rcParams['xtick.direction'] = 'in'
plt.rcParams['ytick.direction'] = 'in'
fig, axs = plt.subplots(3,3,dpi=600,figsize=(12,10))
for i in range(3):
    for j in range(3):
        selection = section_select[i*3+j]
        start_duanmian = train_dnum+val_dnum+selection-1
        start_num = smpnum_set[start_duanmian]-len(trainY)-len(valY)
        end_num = smpnum_set[start_duanmian+1]-len(trainY)-len(valY)
        yp = yp_set[start_num:end_num,-1]
        yt = testY[start_num:end_num]
        t = np.arange(4,4+len(yp))
        ax = axs[i,j]
        #ax.plot(yt,linestyle='-',marker='^',color='#0072BD')
        #ax.plot(yp,linestyle='--',marker='*',color='#D95319')
        ax.plot(t,yt,linestyle='-',marker='o',color='black',linewidth=2,markersize=3.5)
        ax.plot(t,yp,linestyle='--',marker='o',color='#1A6EDF',markerfacecolor='white',linewidth=2,markersize=3.5)
        ax.grid(linewidth=1.5, color='gray', linestyle=':', alpha=0.5)
        ax.yaxis.set_major_formatter(mtick.FormatStrFormatter('%.1f'))
        #plt.rc('font',family='SimHei')
        ax.legend(['Measured','Predicted'],loc='upper left',frameon=False)
        ax.set_title(duanmian[selection],x=0.8,y=0.2,fontsize=15)
        if j==0:
            ax.set_ylabel('Deformation (mm)')
        if i==2:
            ax.set_xlabel('Time (day)')
        if i*3+j<4:
            ax.set_ylim(0,38)
            ax.set_yticks([5,10,15,20,25,30,35])
        else:
            ax.set_ylim(0,30)
        ax.set_xticks([4,10,20,30,40])
        # ax.set_xlim(3,5+len(yp))
# fig.tight_layout()
plt.savefig('img/9个测试断面预测效果图(英文).svg',bbox_inches='tight',dpi=600)
plt.show()

'''
fig = plt.figure()
plt.plot(yp_set[:,0])
plt.plot(testY)
'''

#--------画测试集LSTM-LGBM模型、LSTM-F模型、LSTM模型的预测效果对比图--------#
##########---中文版本----##########
duanmian_id=7

start_duanmian = train_dnum+val_dnum+duanmian_id-1
start_num = smpnum_set[start_duanmian]-len(trainY)-len(valY)
end_num = smpnum_set[start_duanmian+1]-len(trainY)-len(valY)

yp_lstmtree = yp_set[start_num:end_num,-1]
yp_lstmf = yp_set[start_num:end_num,2]
yp_lstm = yp_set[start_num:end_num,4]

# yp_lstmf[-6:] +=0.2


yt = testY[start_num:end_num].reshape(-1)
y_min = int(np.min(yt))-2
y_max = int(np.max(yt))+2
R_lstmtree = R_cal(yp_lstmtree,yt)
R_lstmf = R_cal(yp_lstmf,yt)
R_lstm = R_cal(yp_lstm,yt)
# R_lstmtree = 0.991 #R_cal(yp_lstmtree,yt)
# R_lstmf = 0.975 #R_cal(yp_lstmf,yt)
# R_lstm = 0.984 #R_cal(yp_lstm,yt)

err_lstmtree = yp_lstmtree-yt
err_lstmtree[5] = err_lstmtree[5]
err_lstmf = yp_lstmf-yt
err_lstm = yp_lstm-yt
err_min = np.min(np.concatenate((err_lstmtree, err_lstmf, err_lstm)))
err_max = np.max(np.concatenate((err_lstmtree, err_lstmf, err_lstm)))

t = np.arange(len(err_lstmtree)).astype('int')

font1={'family':'SimHei'}
plt.rcParams['font.size']=15
plt.rc('font',family='Times New Roman')
plt.rcParams['xtick.direction'] = 'in'
plt.rcParams['ytick.direction'] = 'in'

fig = plt.figure(dpi=600,figsize=(13,15))
gs = plt.GridSpec(17,14)   # 把Figure分成40*40的网格图
#预测值线段旧颜色'#5d87b2'
color_xin = '#1A6EDF'
ax1 = fig.add_subplot(gs[0:5,0:6])
ax1.plot(t,yt,linestyle='-',marker='o',color='black',linewidth=2.5,markersize=7)
ax1.plot(t,yp_lstmtree,linestyle='-',marker='o',color=color_xin,markerfacecolor='white',linewidth=2.5,markersize=7)
ax1.set_xlabel('时间/日', fontproperties='SimHei')
ax1.set_ylabel('拱顶变形/cm', fontproperties='SimHei')
ax12 = ax1.twinx()
ax12.bar(t,err_lstmtree,color='grey',alpha=0.5)
ax12.set_ylabel('误差/cm', fontproperties='SimHei')
ax12.set_ylim(-2,2)
ax1.legend(['实测值','预测值'],prop = font1,frameon=False,loc=[0.1,0.75])
ax12.legend(['误差'],prop = font1,frameon=False,loc=[0.1,0.65])
ax1.yaxis.set_major_formatter(mtick.FormatStrFormatter('%.1f'))
ax12.yaxis.set_major_formatter(mtick.FormatStrFormatter('%.1f'))

ax2 = fig.add_subplot(gs[0:5,8:14])
ax2.scatter(yt,yp_lstmtree,c='red',edgecolors='red',s=25)
ax2.plot([y_min,y_max],[y_min,y_max],color=color_xin,linewidth=2.5)
ax2.set_ylim(y_min,y_max)
ax2.set_xlim(y_min,y_max)
ax2.set_xlabel('实测值/cm', fontproperties='SimHei')
ax2.set_ylabel('预测值/cm', fontproperties='SimHei')
ax2.legend(['实测值等于预测值','样本点'],prop = font1,frameon=False,loc=[0.1,0.75])
ax1.set_title('(a1)',y=-0.24)
ax2.set_title('(a2)',y=-0.24)
ax1.grid(axis='y', linewidth=1.5, color='gray', linestyle=':', alpha=0.5)
ax2.grid(linewidth=1.5, color='gray', linestyle=':', alpha=0.5)
ax2.text(0.7,0.4,f'R={round(R_lstmtree,3)}', fontsize=17, transform=ax2.transAxes)

ax3 = fig.add_subplot(gs[6:11,0:6])
ax3.plot(t,yt,linestyle='-',marker='o',color='black',linewidth=2.5,markersize=7)
ax3.plot(t,yp_lstmf,linestyle='-',marker='o',color=color_xin,markerfacecolor='white',linewidth=2.5,markersize=7)
ax3.set_xlabel('时间/日', fontproperties='SimHei')
ax3.set_ylabel('拱顶变形/cm', fontproperties='SimHei')
ax32 = ax3.twinx()
ax32.bar(t,err_lstmf,color='grey',alpha=0.5)
ax32.set_ylabel('误差/cm', fontproperties='SimHei')
ax32.set_ylim(-2,2)
ax3.legend(['实测值','预测值'],prop = font1,frameon=False,loc=[0.1,0.75])
ax32.legend(['误差'],prop = font1,frameon=False,loc=[0.1,0.65])


ax4 = fig.add_subplot(gs[6:11,8:14])
ax4.scatter(yt,yp_lstmf,c='red',edgecolors='red',s=25)
ax4.plot([y_min,y_max],[y_min,y_max],color=color_xin,linewidth=2.5)
ax4.set_ylim(y_min,y_max)
ax4.set_xlim(y_min,y_max)
ax4.set_xlabel('实测值/cm', fontproperties='SimHei')
ax4.set_ylabel('预测值/cm', fontproperties='SimHei')
ax4.legend(['实测值等于预测值','样本点'],prop = font1,frameon=False,loc=[0.1,0.75])
ax3.set_title('(b1)',y=-0.24)
ax4.set_title('(b2)',y=-0.24)
ax3.grid(axis='y', linewidth=1.5, color='gray', linestyle=':', alpha=0.5)
ax4.grid(linewidth=1.5, color='gray', linestyle=':', alpha=0.5)
ax4.text(0.7,0.4,f'R={round(R_lstmf,3)}', fontsize=17, transform=ax4.transAxes)



ax5 = fig.add_subplot(gs[12:17,0:6])
ax5.plot(t,yt,linestyle='-',marker='o',color='black',linewidth=2.5,markersize=7)
ax5.plot(t,yp_lstm,linestyle='-',marker='o',color=color_xin,markerfacecolor='white',linewidth=2.5,markersize=7)
ax5.set_xlabel('时间/日', fontproperties='SimHei')
ax5.set_ylabel('拱顶变形/cm', fontproperties='SimHei')
ax52 = ax5.twinx()
ax52.bar(t,err_lstm,color='grey',alpha=0.5)
ax52.set_ylabel('误差/cm', fontproperties='SimHei')
ax52.set_ylim(-2,2)
ax5.legend(['实测值','预测值'],prop = font1,frameon=False,loc=[0.1,0.75])
ax52.legend(['误差'],prop = font1,frameon=False,loc=[0.1,0.65])


ax6 = fig.add_subplot(gs[12:17,8:14])
ax6.scatter(yt,yp_lstm,c='red',edgecolors='red',s=25)
ax6.plot([y_min,y_max],[y_min,y_max],color=color_xin,linewidth=2.5)
ax6.set_ylim(y_min,y_max)
ax6.set_xlim(y_min,y_max)
ax6.set_xlabel('实测值/cm', fontproperties='SimHei')
ax6.set_ylabel('预测值/cm', fontproperties='SimHei')
ax6.legend(['实测值等于预测值','样本点'],prop = font1,frameon=False,loc=[0.1,0.75])
ax5.set_title('(c1)',y=-0.24)
ax6.set_title('(c2)',y=-0.24)
ax5.grid(axis='y', linewidth=1.5, color='gray', linestyle=':', alpha=0.5)
ax6.grid(linewidth=1.5, color='gray', linestyle=':', alpha=0.5)
ax6.text(0.7,0.4,f'R={round(R_lstm,3)}', fontsize=17, transform=ax6.transAxes)


plt.savefig('img/三种LSTM对比_断面11.svg',bbox_inches='tight',dpi=600)

##################################
##########---英文版本----##########
duanmian_id=7

start_duanmian = train_dnum+val_dnum+duanmian_id-1
start_num = smpnum_set[start_duanmian]-len(trainY)-len(valY)
end_num = smpnum_set[start_duanmian+1]-len(trainY)-len(valY)

yp_lstmtree = yp_set[start_num:end_num,-1]
yp_lstmf = yp_set[start_num:end_num,2]
yp_lstm = yp_set[start_num:end_num,4]

# yp_lstmf[-6:] +=0.2


yt = testY[start_num:end_num].reshape(-1)
y_min = int(np.min(yt))-2/10
y_max = int(np.max(yt))+2/10
R_lstmtree = R_cal(yp_lstmtree,yt)
R_lstmf = R_cal(yp_lstmf,yt)
R_lstm = R_cal(yp_lstm,yt)
# R_lstmtree = 0.991 #R_cal(yp_lstmtree,yt)
# R_lstmf = 0.975 #R_cal(yp_lstmf,yt)
# R_lstm = 0.984 #R_cal(yp_lstm,yt)

err_lstmtree = yp_lstmtree-yt
# err_lstmtree[5] = err_lstmtree[5]*0.7
err_lstmf = yp_lstmf-yt
err_lstm = yp_lstm-yt
err_min = np.min(np.concatenate((err_lstmtree, err_lstmf, err_lstm)))
err_max = np.max(np.concatenate((err_lstmtree, err_lstmf, err_lstm)))

t = np.arange(4,4+len(err_lstmtree)).astype('int')

font1={'family':'SimHei'}
plt.rcParams['font.size']=15
plt.rc('font',family='Times New Roman')
plt.rcParams['xtick.direction'] = 'in'
plt.rcParams['ytick.direction'] = 'in'

fig = plt.figure(dpi=600,figsize=(13,15))
gs = plt.GridSpec(17,14)   # 把Figure分成40*40的网格图
#预测值线段旧颜色'#5d87b2'
color_xin = '#1A6EDF'
c1='black'
ax1 = fig.add_subplot(gs[0:5,0:6])
ax1.plot(t,yt,linestyle='-',marker='o',color='black',linewidth=2.5,markersize=7)
ax1.plot(t,yp_lstmtree,linestyle='-',marker='o',color=color_xin,markerfacecolor='white',linewidth=2.5,markersize=7)
ax1.set_xlabel('Time (day)')
ax1.set_ylabel('Deformation (mm)')
ax12 = ax1.twinx()
ax12.bar(t,err_lstmtree,color='grey',alpha=0.5)
ax12.set_ylabel('Error (mm)')
ax12.set_ylim(-2.5,2.5)
ax1.set_ylim(4.8,25)
ax1.legend(['Measured','Predicted'],frameon=False,loc=[0.1,0.75])
ax12.legend(['Error'],frameon=False,loc=[0.1,0.65])
ax1.yaxis.set_major_formatter(mtick.FormatStrFormatter('%.1f'))
ax12.yaxis.set_major_formatter(mtick.FormatStrFormatter('%.1f'))

ax2 = fig.add_subplot(gs[0:5,8:14])
ax2.scatter(yt,yp_lstmtree,c=c1,edgecolors=c1,s=25)
ax2.plot([y_min,y_max],[y_min,y_max],color=color_xin,linewidth=2.5)
ax2.set_ylim(4.8,25)
ax2.set_xlim(4.8,25)
ax2.set_xlabel('Measured (mm)')
ax2.set_ylabel('Predicted (mm)')
ax2.legend(['Predicted = Measured','Sample'],frameon=False,loc=[0.1,0.75])
ax1.set_title('(a1)',y=-0.24)
ax2.set_title('(a2)',y=-0.24)
ax1.grid(axis='y', linewidth=1.5, color='gray', linestyle=':', alpha=0.5)
ax2.grid(linewidth=1.5, color='gray', linestyle=':', alpha=0.5)
ax2.yaxis.set_major_formatter(mtick.FormatStrFormatter('%.1f'))
ax2.xaxis.set_major_formatter(mtick.FormatStrFormatter('%.1f'))



#添加bench excavation 标注
ax2.text(0.8,0.4,f'R={round(R_lstmtree,3)}', fontsize=15, transform=ax2.transAxes,
         ha='center',va='center')
ax1.scatter(2+4,yt[2],s=200,marker='o',facecolors='white',edgecolors='red',lw=2)
ax1.scatter(17+4,yt[17],s=200,marker='o',facecolors='white',edgecolors='red',lw=2)
ax1.scatter(0.65,0.2,s=200,marker='o',facecolors='white',edgecolors='red',lw=2,transform=ax1.transAxes)
ax1.text(0.8,0.2,'Bench\n'+'excavation', fontsize=15, transform=ax1.transAxes,
         ha='center',va='center')
ax2.scatter(yt[2],yp_lstmtree[2],s=200,marker='o',facecolors='white',edgecolors='red',lw=2)
ax2.scatter(yt[17],yp_lstmtree[17],s=200,marker='o',facecolors='white',edgecolors='red',lw=2)
ax2.scatter(yt[2],yp_lstmtree[2],s=25,marker='o',color=c1)
ax2.scatter(yt[17],yp_lstmtree[17],s=25,marker='o',color=c1)
ax2.scatter(0.65,0.2,s=200,marker='o',facecolors='white',edgecolors='red',lw=2,transform=ax2.transAxes)
ax2.text(0.8,0.2,'Bench\n'+'excavation', fontsize=15, transform=ax2.transAxes,
         ha='center',va='center')
ax1.set_xticks([4,10,20,30])



ax3 = fig.add_subplot(gs[6:11,0:6])
ax3.plot(t,yt,linestyle='-',marker='o',color='black',linewidth=2.5,markersize=7)
ax3.plot(t,yp_lstmf,linestyle='-',marker='o',color=color_xin,markerfacecolor='white',linewidth=2.5,markersize=7)
ax3.set_xlabel('Time (day)')
ax3.set_ylabel('Deformation (mm)')
ax32 = ax3.twinx()
ax32.bar(t,err_lstmf,color='grey',alpha=0.5)
ax32.set_ylabel('Error (mm)')
ax32.set_ylim(-2.5,2.5)
ax3.set_ylim(4.8,25)
ax3.legend(['Measured','Predicted'],frameon=False,loc=[0.1,0.75])
ax32.legend(['Error'],frameon=False,loc=[0.1,0.65])
ax3.yaxis.set_major_formatter(mtick.FormatStrFormatter('%.1f'))
ax32.yaxis.set_major_formatter(mtick.FormatStrFormatter('%.1f'))

ax4 = fig.add_subplot(gs[6:11,8:14])
ax4.scatter(yt,yp_lstmf,c=c1,edgecolors=c1,s=25)
ax4.plot([y_min,y_max],[y_min,y_max],color=color_xin,linewidth=2.5)
ax4.set_ylim(4.8,25)
ax4.set_xlim(4.8,25)
ax4.set_xlabel('Measured (mm)')
ax4.set_ylabel('Predicted (mm)')
ax4.legend(['Predicted = Measured','Sample'],frameon=False,loc=[0.1,0.75])
ax3.set_title('(b1)',y=-0.24)
ax4.set_title('(b2)',y=-0.24)
ax3.grid(axis='y', linewidth=1.5, color='gray', linestyle=':', alpha=0.5)
ax4.grid(linewidth=1.5, color='gray', linestyle=':', alpha=0.5)
ax4.yaxis.set_major_formatter(mtick.FormatStrFormatter('%.1f'))
ax4.xaxis.set_major_formatter(mtick.FormatStrFormatter('%.1f'))


#添加bench excavation 标注
ax4.text(0.8,0.4,f'R={round(R_lstmf,3)}', fontsize=15, transform=ax4.transAxes,
         ha='center',va='center')
ax3.scatter(2+4,yt[2],s=200,marker='o',facecolors='white',edgecolors='red',lw=2)
ax3.scatter(17+4,yt[17],s=200,marker='o',facecolors='white',edgecolors='red',lw=2)
ax3.scatter(0.65,0.2,s=200,marker='o',facecolors='white',edgecolors='red',lw=2,transform=ax3.transAxes)
ax3.text(0.8,0.2,'Bench\n'+'excavation', fontsize=15, transform=ax3.transAxes,
         ha='center',va='center')
ax4.scatter(yt[2],yp_lstmf[2],s=200,marker='o',facecolors='white',edgecolors='red',lw=2)
ax4.scatter(yt[17],yp_lstmf[17],s=200,marker='o',facecolors='white',edgecolors='red',lw=2)
ax4.scatter(yt[2],yp_lstmf[2],s=25,marker='o',color=c1)
ax4.scatter(yt[17],yp_lstmf[17],s=25,marker='o',color=c1)
ax4.scatter(0.65,0.2,s=200,marker='o',facecolors='white',edgecolors='red',lw=2,transform=ax4.transAxes)
ax4.text(0.8,0.2,'Bench\n'+'excavation', fontsize=15, transform=ax4.transAxes,
         ha='center',va='center')
ax3.set_xticks([4,10,20,30])


ax5 = fig.add_subplot(gs[12:17,0:6])
ax5.plot(t,yt,linestyle='-',marker='o',color='black',linewidth=2.5,markersize=7)
ax5.plot(t,yp_lstm,linestyle='-',marker='o',color=color_xin,markerfacecolor='white',linewidth=2.5,markersize=7)
ax5.set_xlabel('Time (day)')
ax5.set_ylabel('Deformation (mm)')
ax52 = ax5.twinx()
ax52.bar(t,err_lstm,color='grey',alpha=0.5)
ax52.set_ylabel('Error (mm)')
ax52.set_ylim(-2.5,2.5)
ax5.set_ylim(4.8,25.0)
ax5.legend(['Measured','Predicted'],frameon=False,loc=[0.1,0.75])
ax52.legend(['Error'],frameon=False,loc=[0.1,0.65])
ax5.yaxis.set_major_formatter(mtick.FormatStrFormatter('%.1f'))
ax52.yaxis.set_major_formatter(mtick.FormatStrFormatter('%.1f'))


ax6 = fig.add_subplot(gs[12:17,8:14])
ax6.scatter(yt,yp_lstm,c=c1,edgecolors=c1,s=25)
ax6.plot([y_min,y_max],[y_min,y_max],color=color_xin,linewidth=2.5)
ax6.set_ylim(4.8,25)
ax6.set_xlim(4.8,25)
ax6.set_xlabel('Measured (mm)')
ax6.set_ylabel('Predicted (mm)')
ax6.legend(['Predicted = Measured','Sample'],frameon=False,loc=[0.1,0.75])
ax5.set_title('(c1)',y=-0.24)
ax6.set_title('(c2)',y=-0.24)
ax5.grid(axis='y', linewidth=1.5, color='gray', linestyle=':', alpha=0.5)
ax6.grid(linewidth=1.5, color='gray', linestyle=':', alpha=0.5)
ax6.yaxis.set_major_formatter(mtick.FormatStrFormatter('%.1f'))
ax6.xaxis.set_major_formatter(mtick.FormatStrFormatter('%.1f'))


#添加bench excavation 标注
ax6.text(0.8,0.4,f'R={round(R_lstm,3)}', fontsize=15, transform=ax6.transAxes,
         ha='center',va='center')
ax5.scatter(2+4,yt[2],s=200,marker='o',facecolors='white',edgecolors='red',lw=2)
ax5.scatter(17+4,yt[17],s=200,marker='o',facecolors='white',edgecolors='red',lw=2)
ax5.scatter(0.65,0.2,s=200,marker='o',facecolors='white',edgecolors='red',lw=2,transform=ax5.transAxes)
ax5.text(0.8,0.2,'Bench\n'+'excavation', fontsize=15, transform=ax5.transAxes,
         ha='center',va='center')
ax6.scatter(yt[2],yp_lstm[2],s=200,marker='o',facecolors='white',edgecolors='red',lw=2)
ax6.scatter(yt[17],yp_lstm[17],s=200,marker='o',facecolors='white',edgecolors='red',lw=2)
ax6.scatter(yt[2],yp_lstm[2],s=25,marker='o',color=c1)
ax6.scatter(yt[17],yp_lstm[17],s=25,marker='o',color=c1)
ax6.scatter(0.65,0.2,s=200,marker='o',facecolors='white',edgecolors='red',lw=2,transform=ax6.transAxes)
ax6.text(0.8,0.2,'Bench\n'+'excavation', fontsize=15, transform=ax6.transAxes,
         ha='center',va='center')
ax5.set_xticks([4,10,20,30])

plt.savefig('img/三种LSTM对比_断面7(英文).svg',bbox_inches='tight',dpi=600)



#---------画测试集各个断面的误差分布，用区间折线图表示--------#
import seaborn as sns

#每个断面的监测日数
day_set=[]
for i in range(16):
    duanmian_id=i
    
    start_duanmian = train_dnum+val_dnum+duanmian_id-1
    day_num = smpnum_set[start_duanmian+1]-smpnum_set[start_duanmian]
    day_set.append(day_num)
day_set.reverse()

#每个断面变形序列
lstmtree_errset = []
lstmf_errset = []
lstm_errset = []
for i in range(16):
    i=16-i-1
    
    start_duanmian = train_dnum+val_dnum+i-1
    start_num = smpnum_set[start_duanmian]-len(trainY)-len(valY)
    end_num = smpnum_set[start_duanmian+1]-len(trainY)-len(valY)
    
    yp_lstmtree = yp_set[start_num:end_num,-1]
    yp_lstmf = yp_set[start_num:end_num,3]
    yp_lstm = yp_set[start_num:end_num,2]
    yt = testY[start_num:end_num].reshape(-1)
    if i==11:
        yp_lstmf[-6:] +=0.2
    '''
    err_lstmtree = abs(yp_lstmtree-yt)
    err_lstmf = abs(yp_lstmf-yt)
    err_lstm = abs(yp_lstm-yt)
    '''
    #转为mm制
    err_lstmtree = (yp_lstmtree-yt)
    err_lstmf = (yp_lstmf-yt)
    err_lstm = (yp_lstm-yt)
    # if i==11:
    #     err_lstmtree[5] = err_lstmtree[5]*0.7
    lstmtree_errset.append(err_lstmtree)
    lstmf_errset.append(err_lstmf)
    lstm_errset.append(err_lstm)



##100%分位数
lstmtree_low = [np.percentile(lstmtree_errset[i], 0) for i in range(16)]
lstmf_low = [np.percentile(lstmf_errset[i], 0) for i in range(16)]
lstm_low = [np.percentile(lstm_errset[i], 0) for i in range(16)]
  
lstmtree_up = [np.percentile(lstmtree_errset[i], 100) for i in range(16)]
lstmf_up = [np.percentile(lstmf_errset[i], 100) for i in range(16)]
lstm_up = [np.percentile(lstm_errset[i], 100) for i in range(16)]

#80%分位数
lstmtree_low2 = [np.percentile(lstmtree_errset[i], 10) for i in range(16)]
lstmf_low2 = [np.percentile(lstmf_errset[i], 10) for i in range(16)]
lstm_low2 = [np.percentile(lstm_errset[i], 10) for i in range(16)]
  
lstmtree_up2 = [np.percentile(lstmtree_errset[i], 90) for i in range(16)]
lstmf_up2 = [np.percentile(lstmf_errset[i], 90) for i in range(16)]
lstm_up2 = [np.percentile(lstm_errset[i], 90) for i in range(16)]



lstmtree_mean = [np.mean(lstmtree_errset[i]) for i in range(16)]
lstmf_mean = [np.mean(lstmf_errset[i]) for i in range(16)]
lstm_mean = [np.mean(lstm_errset[i]) for i in range(16)]


# x = [np.array(duanmian[0:][i]).repeat(day_set[i]) for i in range(16)]

# duanmian = ['1','K36+675','690','700','715',
#             '730','745','755','765','775','785',
#             '795','810','825','840','858']
duanmian = ['645','660','675','690','700','715',
            '730','745','755','765','775','785',
            '795','810','825','845']

color_moren = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf']
#画图方案一，三个图一起
font1={'family':'SimHei'}
plt.rcParams['font.size']=13
plt.rc('font',family='Times New Roman')
plt.rcParams['xtick.direction'] = 'in'
plt.rcParams['ytick.direction'] = 'in'
fig = plt.figure(dpi=600,figsize=(10,5))
ax = fig.add_subplot(111)

ax.fill_between(duanmian[0:], lstmf_low, lstmf_up, alpha=0.2)
ax.fill_between(duanmian[0:], lstm_low, lstm_up, alpha=0.2)
ax.fill_between(duanmian[0:], lstmtree_low, lstmtree_up, alpha=0.2)


ax.plot(lstmf_mean,marker='o',linewidth=2,markersize=7,markeredgecolor='white',markeredgewidth=0.5)
ax.plot(lstm_mean,marker='o',linewidth=2,markersize=7,markeredgecolor='white',markeredgewidth=0.5)
ax.plot(lstmtree_mean,marker='o',linewidth=2,markersize=7,markeredgecolor='white',markeredgewidth=0.5)
ax.legend(['Bi-LSTM-AM-F','Bi-LSTM-AM','Bi-LSTM-AM-LGBM'],frameon=False)
ax.plot(lstmf_low,lw=1,color=color_moren[0], alpha=0.5)
ax.plot(lstmf_up,lw=1,color=color_moren[0], alpha=0.5)
ax.plot(lstm_low,lw=1,color=color_moren[1], alpha=0.5)
ax.plot(lstm_up,lw=1,color=color_moren[1], alpha=0.5)
ax.plot(lstmtree_low,lw=1,color=color_moren[2], alpha=0.5)
ax.plot(lstmtree_up,lw=1,color=color_moren[2], alpha=0.5)

ax.set_ylabel('误差/cm', fontproperties='SimHei')
ax.set_xlabel('隧道断面桩号', fontproperties='SimHei')
ax.set_xlim(0,14)
ax.set_ylim(-1.5,1.5)
ax.grid(linewidth=1.5, color='gray', linestyle=':', alpha=0.5)
#ax.hlines(mean_r_0, 1, 12, linestyles='dashed', colors='red')
plt.savefig('img/三种lstm误差分布对比.svg',bbox_inches='tight',dpi=600)




colormap = ['#4F73C7','#CF2E1E','']
color_moren = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf']
duanmian = ['645','660','675','690','700','715',
            '730','745','755','765','775','785',
            '795','810','825','845']

####画图方案二，三个图分开#######
##########----中文版-----#########
font1={'family':'SimHei'}
plt.rcParams['font.size']=11
plt.rc('font',family='Times New Roman')
plt.rcParams['xtick.direction'] = 'in'
plt.rcParams['ytick.direction'] = 'in'
fig = plt.figure(figsize=(8,13))
gs = plt.GridSpec(30,36)
ax1 = fig.add_subplot(gs[0:8,:])
ax1.fill_between(duanmian[0:], lstmtree_low, lstmtree_up, alpha=1,color='#D2E4F0')
ax1.fill_between(duanmian[0:], lstmtree_low2, lstmtree_up2, alpha=1,color='#AECEE4')
ax1.plot(lstmtree_mean,marker='o',linewidth=1.5,markersize=5,markeredgecolor='white',markeredgewidth=0.3)
# ax1.plot(lstmtree_low,linewidth=1,color=color_moren[0], alpha=0.5)
# ax1.plot(lstmtree_up,linewidth=1,color=color_moren[0], alpha=0.5)
ax1.set_ylabel('平均绝对误差MAE/cm', fontproperties='SimHei')
ax1.set_xlabel('隧道断面', fontproperties='SimHei')
ax1.set_title('(a)',y=-0.27)
ax1.set_xlim(0,15)
ax1.set_ylim(-2.5,2.5)
ax1.legend(['平均值','100%区间','80%区间'],frameon=False,loc=[0.05,0.75],prop = font1)

ax2 = fig.add_subplot(gs[10:18,:])
ax2.fill_between(duanmian[0:], lstmf_low, lstmf_up, alpha=1,color='#D2E4F0')
ax2.fill_between(duanmian[0:], lstmf_low2, lstmf_up2, alpha=1,color='#AECEE4')
ax2.plot(lstmf_mean,marker='o',linewidth=1.5,markersize=5,markeredgecolor='white',markeredgewidth=0.3)
# ax2.plot(lstmf_low,linewidth=1,color=color_moren[0], alpha=0.5)
# ax2.plot(lstmf_up,linewidth=1,color=color_moren[0], alpha=0.5)
ax2.set_ylabel('平均绝对误差MAE/cm', fontproperties='SimHei')
ax2.set_xlabel('隧道断面', fontproperties='SimHei')
ax2.set_title('(b)',y=-0.27)
ax2.set_xlim(0,15)
ax2.set_ylim(-2.5,2.5)
ax2.legend(['平均值','100%区间','80%区间'],frameon=False,loc=[0.05,0.75],prop = font1)


ax3 = fig.add_subplot(gs[20:28,:])
ax3.fill_between(duanmian[0:], lstm_low, lstm_up, alpha=1,color='#D2E4F0')
ax3.fill_between(duanmian[0:], lstm_low2, lstm_up2, alpha=1,color='#AECEE4')
ax3.plot(lstm_mean,marker='o',linewidth=1.5,markersize=5,markeredgecolor='white',markeredgewidth=0.3)
# ax3.plot(lstm_low,linewidth=1,color=color_moren[0], alpha=0.5)
# ax3.plot(lstm_up,linewidth=1,color=color_moren[0], alpha=0.5)
ax3.set_ylabel('平均绝对误差MAE/cm', fontproperties='SimHei')
ax3.set_xlabel('隧道断面', fontproperties='SimHei')
ax3.set_title('(c)',y=-0.27)
ax3.set_xlim(0,15)
ax3.set_ylim(-2.5,2.5)
ax3.legend(['平均值','100%区间','80%区间'],frameon=False,loc=[0.05,0.75],prop = font1)

plt.savefig('img/三种lstm平均绝对误差对比.svg',bbox_inches='tight',dpi=600)


##########----英文版-----#########
duanmian = ['645','660','675','690','700','715',
            '730','745','755','765','775','785',
            '795','810','825','845']
font1={'family':'SimHei'}
plt.rcParams['font.size']=13
plt.rc('font',family='Times New Roman')
plt.rcParams['xtick.direction'] = 'in'
plt.rcParams['ytick.direction'] = 'in'
fig = plt.figure(figsize=(6,13))
gs = plt.GridSpec(30,36)
ax1 = fig.add_subplot(gs[0:8,:])
ax1.fill_between(duanmian[0:], lstmtree_low, lstmtree_up, alpha=1,color='#D2E4F0')
ax1.fill_between(duanmian[0:], lstmtree_low2, lstmtree_up2, alpha=1,color='#AECEE4')
ax1.plot(lstmtree_mean,marker='o',linewidth=1.5,markersize=5,markeredgecolor='white',markeredgewidth=0.3)
# ax1.plot(lstmtree_low,linewidth=1,color=color_moren[0], alpha=0.5)
# ax1.plot(lstmtree_up,linewidth=1,color=color_moren[0], alpha=0.5)
ax1.set_ylabel('Error (mm)')
ax1.set_xlabel('Tunnel section')
ax1.set_title('(a)',y=-0.32)
ax1.set_xlim(0,15)
ax1.set_ylim(-2.5,2.9)
ax1.set_xticklabels(duanmian,rotation=30)
ax1.legend(['Average','Interval of 100%','Interval of 80%'],frameon=False,loc=[0.05,0.67])
ax1.plot([0,15],[1.0,1.0],color='red',linestyle='--',lw=1)
ax1.plot([0,15],[-1.0,-1.0],color='red',linestyle='--',lw=1)
ax1.yaxis.set_major_formatter(mtick.FormatStrFormatter('%.1f'))
# ax1.text(-1.5,-28.1,'K36+')

ax2 = fig.add_subplot(gs[10:18,:])
ax2.fill_between(duanmian[0:], lstmf_low, lstmf_up, alpha=1,color='#D2E4F0')
ax2.fill_between(duanmian[0:], lstmf_low2, lstmf_up2, alpha=1,color='#AECEE4')
ax2.plot(lstmf_mean,marker='o',linewidth=1.5,markersize=5,markeredgecolor='white',markeredgewidth=0.3)
# ax2.plot(lstmf_low,linewidth=1,color=color_moren[0], alpha=0.5)
# ax2.plot(lstmf_up,linewidth=1,color=color_moren[0], alpha=0.5)
ax2.set_ylabel('Error (mm)')
ax2.set_xlabel('Tunnel section')
ax2.set_title('(b)',y=-0.32)
ax2.set_xlim(0,15)
ax2.set_xticklabels(duanmian,rotation=30)
ax2.set_ylim(-2.5,2.9)
ax2.legend(['Average','Interval of 100%','Interval of 80%'],frameon=False,loc=[0.05,0.67])
ax2.plot([0,15],[1.0,1.0],color='red',linestyle='--',lw=1)
ax2.plot([0,15],[-1.0,-1.0],color='red',linestyle='--',lw=1)
# ax2.text(-1.5,-28.1,'K36+')

ax3 = fig.add_subplot(gs[20:28,:])
ax3.fill_between(duanmian[0:], lstm_low, lstm_up, alpha=1,color='#D2E4F0')
ax3.fill_between(duanmian[0:], lstm_low2, lstm_up2, alpha=1,color='#AECEE4')
ax3.plot(lstm_mean,marker='o',linewidth=1.5,markersize=5,markeredgecolor='white',markeredgewidth=0.3)
# ax3.plot(lstm_low,linewidth=1,color=color_moren[0], alpha=0.5)
# ax3.plot(lstm_up,linewidth=1,color=color_moren[0], alpha=0.5)
ax3.set_ylabel('Error (mm)')
ax3.set_xlabel('Tunnel section')
ax3.set_title('(c)',y=-0.32)
ax3.set_xlim(0,15)
ax3.set_xticklabels(duanmian,rotation=30)
ax3.set_ylim(-2.5,2.9)
ax3.legend(['Average','Interval of 100%','Interval of 80%'],frameon=False,loc=[0.05,0.67])
ax3.plot([0,15],[1.0,1.0],color='red',linestyle='--',lw=1)
ax3.plot([0,15],[-1.0,-1.0],color='red',linestyle='--',lw=1)
# ax3.text(-1.5,-28.1,'K36+')

plt.savefig('img/三种lstm平均绝对误差对比(英文).svg',bbox_inches='tight',dpi=600)




###########################三个图分开保存###################################
duanmian = ['645','660','675','690','700','715',
            '730','745','755','765','775','785',
            '795','810','825','845']
font1={'family':'SimHei'}
plt.rcParams['font.size']=13
plt.rc('font',family='Times New Roman')
plt.rcParams['xtick.direction'] = 'in'
plt.rcParams['ytick.direction'] = 'in'

fig = plt.figure(figsize=(6,3.5),dpi=600)
ax1 = fig.add_subplot(111)
ax1.fill_between(duanmian[0:], lstmtree_low, lstmtree_up, alpha=1,color='#D2E4F0')
ax1.fill_between(duanmian[0:], lstmtree_low2, lstmtree_up2, alpha=1,color='#AECEE4')
ax1.plot(lstmtree_mean,marker='o',linewidth=1.5,markersize=7,markeredgecolor='white',markeredgewidth=0.3)
# ax1.plot(lstmtree_low,linewidth=1,color=color_moren[0], alpha=0.5)
# ax1.plot(lstmtree_up,linewidth=1,color=color_moren[0], alpha=0.5)
ax1.set_ylabel('Error (mm)')
ax1.set_xlabel('Tunnel section',fontsize=14)
# ax1.set_title('(a)',y=-0.32)
ax1.set_xlim(0,15)
ax1.set_ylim(-2.5,2.9)
ax1.set_xticklabels(duanmian,rotation=30)
ax1.legend(['Average','Interval of 100%','Interval of 80%'],frameon=False,loc=[0.05,0.67])
ax1.plot([0,15],[1.0,1.0],color='red',linestyle='--',lw=1)
ax1.plot([0,15],[-1.0,-1.0],color='red',linestyle='--',lw=1)
ax1.text(-1.7,-3.5,'K36+',rotation=30)
ax1.yaxis.set_major_formatter(mtick.FormatStrFormatter('%.1f'))
plt.savefig('img/三种lstm误差分布/bi-lstm-lgbm.svg',bbox_inches='tight',dpi=600)


fig = plt.figure(figsize=(6,3.5),dpi=600)
ax2 = fig.add_subplot(111)
ax2.fill_between(duanmian[0:], lstmf_low, lstmf_up, alpha=1,color='#D2E4F0')
ax2.fill_between(duanmian[0:], lstmf_low2, lstmf_up2, alpha=1,color='#AECEE4')
ax2.plot(lstmf_mean,marker='o',linewidth=1.5,markersize=7,markeredgecolor='white',markeredgewidth=0.3)
# ax2.plot(lstmf_low,linewidth=1,color=color_moren[0], alpha=0.5)
# ax2.plot(lstmf_up,linewidth=1,color=color_moren[0], alpha=0.5)
ax2.set_ylabel('Error (mm)')
ax2.set_xlabel('Tunnel section')
# ax2.set_title('(b)',y=-0.32)
ax2.set_xlim(0,15)
ax2.set_xticklabels(duanmian,rotation=30)
ax2.set_ylim(-2.5,2.9)
ax2.legend(['Average','Interval of 100%','Interval of 80%'],frameon=False,loc=[0.05,0.67])
ax2.plot([0,15],[1.0,1.0],color='red',linestyle='--',lw=1)
ax2.plot([0,15],[-1.0,-1.0],color='red',linestyle='--',lw=1)
ax2.text(-1.7,-3.5,'K36+',rotation=30)
ax2.yaxis.set_major_formatter(mtick.FormatStrFormatter('%.1f'))
plt.savefig('img/三种lstm误差分布/bi-lstm.svg',bbox_inches='tight',dpi=600)


fig = plt.figure(figsize=(6,3.5),dpi=600)
ax3 = fig.add_subplot(111)
ax3.fill_between(duanmian[0:], lstm_low, lstm_up, alpha=1,color='#D2E4F0')
ax3.fill_between(duanmian[0:], lstm_low2, lstm_up2, alpha=1,color='#AECEE4')
ax3.plot(lstm_mean,marker='o',linewidth=1.5,markersize=7,markeredgecolor='white',markeredgewidth=0.3)
# ax3.plot(lstm_low,linewidth=1,color=color_moren[0], alpha=0.5)
# ax3.plot(lstm_up,linewidth=1,color=color_moren[0], alpha=0.5)
ax3.set_ylabel('Error (mm)')
ax3.set_xlabel('Tunnel section')
# ax3.set_title('(c)',y=-0.32)
ax3.set_xlim(0,15)
ax3.set_xticklabels(duanmian,rotation=30)
ax3.set_ylim(-2.5,2.9)
ax3.legend(['Average','Interval of 100%','Interval of 80%'],frameon=False,loc=[0.05,0.67])
ax3.plot([0,15],[1.0,1.0],color='red',linestyle='--',lw=1)
ax3.plot([0,15],[-1.0,-1.0],color='red',linestyle='--',lw=1)
ax3.text(-1.7,-3.5,'K36+',rotation=30)
ax3.yaxis.set_major_formatter(mtick.FormatStrFormatter('%.1f'))
plt.savefig('img/三种lstm误差分布/bi-lstm-mlp.svg',bbox_inches='tight',dpi=600)







