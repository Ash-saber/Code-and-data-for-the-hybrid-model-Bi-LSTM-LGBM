# -*- coding: utf-8 -*-
"""
Created on Fri Apr 14 21:08:48 2023

@author: Steve
"""

import pandas as pd
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
import tensorflow as tf
from xgboost import XGBRegressor
from lightgbm import LGBMRegressor
import os
path = r'D:\Program project\python project\pytorch\time_seris'
os.chdir(path)

from tools.util import *
from tools.model_create import *
from tools.quantile_reg
np.random.seed(42)
tf.random.set_seed(42)

def create_dataset(dataset, Fdataset, look_back, predict_step): #构造数据集结构函数
    #dataset是变形、围岩等级、掌子面里程时间序列，(n,3)的矩阵
    #Fdataset是非时间序列因素，仅与断面有关
    dataX, dataY = [], []                #定义空数组的方式
    dataF = []
    dataF2 = []
    for j in range(0,len(dataset)-look_back-predict_step+1):
        a_array = dataset[j:(j+look_back)]
        dataX.append(a_array)            #这个是所有的data
        b_array = dataset[(j + look_back):j+look_back+predict_step,0]
        dataY.append(b_array)            #这个应该是变形序列
        
        d_array = dataset[(j + look_back):j+look_back+predict_step,1:]
        d_array = np.squeeze(d_array,0)  #这个是围岩等级、掌子面里程序列
        dataF2.append(d_array)
        
        c_array = Fdataset
        dataF.append(c_array)
    return np.array(dataX),np.array(dataY),np.array(dataF),np.array(dataF2)



data_select = 0    #选择使用的数据

if data_select==0:
    file1 = r'所有数据\原始\left_top.xlsx'
    file2 = r'所有数据\原始\right_top.xlsx'
elif data_select==1:
    file1 = r'所有数据\直线\left_top_zhi.xlsx'
    file2 = r'所有数据\直线\right_top_zhi.xlsx'
elif data_select==2:
    file1 = r'所有数据\曲线\left_top_qu.xlsx'
    file2 = r'所有数据\曲线\right_top_qu.xlsx'


LTSdata_x = np.empty(shape=[0,3,2])
LTSdata_y = np.empty(shape=[0,1])

LFdata = np.empty(shape=[0,4])
#Fdata2 = np.empty(shape=[0,2])

wb1 = load_workbook(file1,data_only=True)
ws1 = wb1.sheetnames
####构造数据集结构-左线#####
smp_num=0
smpnum_set = []#样本随断面增加而增加的记录，后面根据断面数量划分数据集要用到
left_dataset = []
for i in range(len(ws1)-2):    #读取ws-2个断面数据，合成数据集
    ######序列数据
    LTSdata = pd.read_excel(file1,sheet_name=i,header=0, usecols = [1,9],engine='openpyxl')#变形、围岩等级、掌子面里程时间序列
    LTSdata = LTSdata.dropna()   #这条命令太重要了
    LTSdata = LTSdata.values
    LTSdata = LTSdata.astype('float32')
    
        
    ####因素数据
    LFactordata = pd.read_excel(file1,sheet_name=i,header=0, usecols = [4,5,10,6],engine='openpyxl')#4个非时间序列因素，只随断面变化
    LFactordata = LFactordata.dropna()   #这条命令太重要了
    LFactordata = LFactordata.values
    LFactordata = LFactordata.astype('float32')
    LFactordata = np.squeeze(LFactordata,0)

    left_dataset.append([LTSdata,LFactordata])

    look_back = 3
    predict_step=1

    X_data,Y_data,F_data,F2_data  = create_dataset(LTSdata,LFactordata, look_back,predict_step)
    print("目前数据集的格式：",X_data.shape,Y_data.shape,F_data.shape,F2_data.shape)
    print('当前样本量',X_data.shape[0])    
    smp_num += X_data.shape[0]
    smpnum_set.append(smp_num)
    print('当前样本量总数',smp_num)
    
    LTSdata_x = np.concatenate((LTSdata_x,X_data),axis = 0)
    LTSdata_y = np.concatenate((LTSdata_y,Y_data),axis = 0)
    
    LFdata = np.concatenate((LFdata,F_data),axis = 0)
    #Fdata2 = np.concatenate((Fdata2,F2_data),axis = 0)

####构造数据集结构-右线#####


RTSdata_x = np.empty(shape=[0,3,2])
RTSdata_y = np.empty(shape=[0,1])

RFdata = np.empty(shape=[0,4])
#RFdata2 = np.empty(shape=[0,2])

wb2 = load_workbook(file2,data_only=True)
ws2 = wb2.sheetnames
right_dataset = []
for k in range(len(ws2)-2):
    ######序列数据
    RTSdata = pd.read_excel(file2,sheet_name=k,header=0, usecols = [1,9])
    RTSdata = RTSdata.dropna()   #这条命令太重要了
    RTSdata = RTSdata.values
    RTSdata = RTSdata.astype('float32')
    
    
    ####因素数据
    RFactordata = pd.read_excel(file2,sheet_name=k,header=0, usecols = [4,5,10,6])
    RFactordata = RFactordata.dropna()   #这条命令太重要了
    RFactordata = RFactordata.values
    RFactordata = RFactordata.astype('float32')
    RFactordata = np.squeeze(RFactordata,0)

    right_dataset.append([RTSdata,RFactordata])

    look_back = 3
    predict_step=1

    X_data,Y_data,F_data,F2_data  = create_dataset(RTSdata,RFactordata, look_back,predict_step)
    print("目前数据集的格式：",X_data.shape,Y_data.shape,F_data.shape,F2_data.shape)
    print('当前样本量（右线）',X_data.shape[0])    
    smp_num += X_data.shape[0]
    smpnum_set.append(smp_num)
    print('当前样本量总数（右线）',smp_num)
    
    RTSdata_x = np.concatenate((RTSdata_x,X_data),axis = 0)
    RTSdata_y = np.concatenate((RTSdata_y,Y_data),axis = 0)
    
    RFdata = np.concatenate((RFdata,F_data),axis = 0)
    #RFdata2 = np.concatenate((RFdata2,F2_data),axis = 0)

####构造总数据集结构#####
full_dataset = left_dataset+right_dataset
TSdata_x = np.concatenate((LTSdata_x,RTSdata_x),axis=0)
#交换Fdata最后一个类别因素与连续因素
Fdata_x = np.concatenate((LFdata,RFdata),axis=0)
Fdata_x[:,[-2,-1]] = Fdata_x[:,[-1,-2]]

TSdata_y = np.concatenate((LTSdata_y,RTSdata_y),axis=0)
ws = ws1[:-2]+ws2[:-2]
duanmian_num = len(ws)#断面总数

#给变形时间序列加高斯噪声
TSdata_yn = np.array([])
for start,end in zip([0]+smpnum_set[:-1],smpnum_set):
    TSdata_yn=np.append(TSdata_yn,add_discrete_noise(TSdata_y[start:end],intensity=2,discrete=0.4))

TSdata_yn = TSdata_yn.reshape(-1,1)
#画个图看下
d_id=11
fig = plt.figure(dpi=600)
plt.plot(TSdata_y[smpnum_set[d_id]:smpnum_set[d_id+1]],marker='o')
plt.plot(TSdata_yn[smpnum_set[d_id]:smpnum_set[d_id+1]],marker='*',linestyle='--')
#plt.plot(TSdata_yn[smpnum_set[5]:smpnum_set[6]].reshape(-1,1)-TSdata_y[smpnum_set[5]:smpnum_set[6]],marker='*',linestyle='--')

#设置数据集划分比例
train_rate = 0.6
val_rate = 0.2
test_rate = 0.2


train_dnum = int(train_rate*duanmian_num)
val_dnum = int(val_rate*duanmian_num)
test_dnum = duanmian_num-train_dnum-val_dnum

trainTS = TSdata_x[:smpnum_set[train_dnum-1]]
trainF = Fdata_x[:smpnum_set[train_dnum-1]]
trainY = TSdata_y[:smpnum_set[train_dnum-1]]


valTS = TSdata_x[smpnum_set[train_dnum-1]:smpnum_set[train_dnum+val_dnum-1]]
valF = Fdata_x[smpnum_set[train_dnum-1]:smpnum_set[train_dnum+val_dnum-1]]
valY = TSdata_y[smpnum_set[train_dnum-1]:smpnum_set[train_dnum+val_dnum-1]]

testTS = TSdata_x[smpnum_set[train_dnum+val_dnum-1]:]
testF = Fdata_x[smpnum_set[train_dnum+val_dnum-1]:]
testY = TSdata_y[smpnum_set[train_dnum+val_dnum-1]:]

#使用自己写的类函数来标准化
Normal_X, Normal_Y = Normalize_timeseris(),Normalize()
Normal_X.fit(trainTS,norm_type=0)
Normal_Y.fit(trainY,norm_type=0)

Normal_licheng = Normalize()
Normal_licheng.fit(trainF[:,-1],norm_type=0)

#写个专门标准化静态因素的函数，前三个为类别变量第四个为连续变量，处理方式不同
def Norm_Fdata(Fdata):
    '''
    Parameters
    ----------
    Fdata  (smp_num,factor_num)
    
    Returns
    -------
    onehot_encoded and normalized Fdata

    '''
    norm_Fdata = np.empty(shape=[Fdata.shape[0],0])
    for i in range(Fdata.shape[1]):
        if i<3:
            norm_Fdata = np.concatenate((norm_Fdata,onehot_encode(Fdata[:,i],Fdata_x[:,i])),axis=1)
        else:
            norm_Fdata = np.concatenate((norm_Fdata,Normal_licheng.norm_transform(Fdata[:,i]).reshape(-1,1)),axis=1)
    return norm_Fdata

#对所有数据进行标准化
trainTS = Normal_X.norm_transform(trainTS)
trainY = Normal_Y.norm_transform(trainY)
trainF = Norm_Fdata(trainF)

valTS = Normal_X.norm_transform(valTS)
valY = Normal_Y.norm_transform(valY)
valF = Norm_Fdata(valF)

testTS = Normal_X.norm_transform(testTS)
#testY = Normal_X.norm_transform(testY)
testF = Norm_Fdata(testF)





















